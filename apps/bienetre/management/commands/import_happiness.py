import os
from django.conf import settings
from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from apps.bienetre.models import HappinessRecord
import pycountry

def get_country_iso(country_name):
    try:
        results = pycountry.countries.search_fuzzy(country_name)
        if results:
            return results[0].alpha_2
    except Exception:
        return None

class Command(BaseCommand):
    help = "Importe les données de bien-être depuis le fichier DataHapiness2025.xlsx (pour 2019 à 2024) en stockant le code ISO du pays."

    def handle(self, *args, **kwargs):
        xlsx_path = os.path.join(settings.BASE_DIR, 'data', 'DataHapiness2025.xlsx')
        self.stdout.write(f"Importation depuis : {xlsx_path}")

        try:
            wb = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
            ws = wb.active

            header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            header_map = {header[i]: i for i in range(len(header))}

            count = 0
            for row in ws.iter_rows(min_row=2):
                try:
                    year = int(row[header_map["Year"]].value)
                except Exception:
                    continue

                if year < 2019 or year > 2024:
                    continue

                try:
                    rank = int(row[header_map["Rank"]].value)
                except Exception:
                    rank = None

                country_cell = row[header_map["Country name"]].value
                if country_cell:
                    country_name = country_cell.strip()
                    iso_code = get_country_iso(country_name)
                    if iso_code:
                        country_iso = iso_code
                    else:
                        country_iso = country_name  # On garde le nom en cas d'échec
                else:
                    country_iso = ""

                def to_float(cell):
                    try:
                        return float(cell.value)
                    except Exception:
                        return None

                ladder_score = to_float(row[header_map["Ladder score"]])
                upperwhisker = to_float(row[header_map["upperwhisker"]])
                lowerwhisker = to_float(row[header_map["lowerwhisker"]])
                log_gdp_per_capita = to_float(row[header_map["Explained by: Log GDP per capita"]])
                social_support = to_float(row[header_map["Explained by: Social support"]])
                healthy_life_expectancy = to_float(row[header_map["Explained by: Healthy life expectancy"]])
                freedom = to_float(row[header_map["Explained by: Freedom to make life choices"]])
                generosity = to_float(row[header_map["Explained by: Generosity"]])
                perceptions = to_float(row[header_map["Explained by: Perceptions of corruption"]])
                dystopia_residual = to_float(row[header_map["Dystopia + residual"]])

                HappinessRecord.objects.create(
                    year=year,
                    rank=rank,
                    country_name=country_iso,
                    ladder_score=ladder_score,
                    upperwhisker=upperwhisker,
                    lowerwhisker=lowerwhisker,
                    log_gdp_per_capita=log_gdp_per_capita,
                    social_support=social_support,
                    healthy_life_expectancy=healthy_life_expectancy,
                    freedom_to_make_life_choices=freedom,
                    generosity=generosity,
                    perceptions_of_corruption=perceptions,
                    dystopia_residual=dystopia_residual,
                )
                count += 1
            self.stdout.write(self.style.SUCCESS(f"Import terminé avec succès. {count} enregistrements créés."))
        except FileNotFoundError:
            self.stderr.write(self.style.ERROR(f"Le fichier {xlsx_path} n'a pas été trouvé."))